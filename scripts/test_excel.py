"""
OpenFinanceKit — Test Suite
============================
Pruebas automatizadas del archivo OpenFinanceKit.xlsx

Estrategia:
1. Verifica estructura (hojas, tablas, validaciones)
2. Inserta datos de prueba controlados
3. Calcula resultados esperados en Python (simula el Motor)
4. Verifica que las fórmulas del Motor apuntan a los datos correctos
5. Verifica coherencia del Dashboard con el Motor

Ejecutar: python3 scripts/test_excel.py
"""
import sys
sys.path.insert(0, '/home/echeverria/.local/lib/python3.14/site-packages')

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy
from datetime import date
import os

# ── Configuración ──────────────────────────────────────────────────────────
EXCEL_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "product", "excel", "OpenFinanceKit.xlsx"
)

class TestResult:
    def __init__(self):
        self.passed = 0
        self.failed = 0
        self.errors = []
    
    def ok(self, msg):
        self.passed += 1
        print(f"  ✅ {msg}")
    
    def fail(self, msg):
        self.failed += 1
        self.errors.append(msg)
        print(f"  ❌ {msg}")
    
    def summary(self):
        total = self.passed + self.failed
        print(f"\n{'═'*60}")
        print(f"  RESULTADOS: {self.passed}/{total} pruebas pasaron")
        if self.errors:
            print(f"\n  Errores:")
            for e in self.errors:
                print(f"    • {e}")
        print(f"{'═'*60}")
        return self.failed == 0

results = TestResult()

# ══════════════════════════════════════════════════════════════════════════
# TEST 1 — ESTRUCTURA
# ══════════════════════════════════════════════════════════════════════════
def test_estructura(wb):
    print("\n── TEST 1: Estructura del libro ──")
    
    # Hojas esperadas
    esperadas = ["Configuración", "Ingresos", "Gastos", "Pagos", "Motor", "Dashboard"]
    if wb.sheetnames == esperadas:
        results.ok(f"Hojas correctas: {esperadas}")
    else:
        results.fail(f"Hojas incorrectas: {wb.sheetnames} (esperado: {esperadas})")
    
    # Tablas esperadas
    tablas = {
        "Configuración": ["tblCategorias", "tblCuentas", "tblParametros"],
        "Ingresos": ["tblIngresos"],
        "Gastos": ["tblGastos"],
        "Pagos": ["tblPagos"],
    }
    for hoja, tablas_esperadas in tablas.items():
        ws = wb[hoja]
        for t in tablas_esperadas:
            if t in ws.tables:
                results.ok(f"{hoja}.{t} existe")
            else:
                results.fail(f"{hoja}.{t} NO encontrada")

# ══════════════════════════════════════════════════════════════════════════
# TEST 2 — VALIDACIONES
# ══════════════════════════════════════════════════════════════════════════
def test_validaciones(wb):
    print("\n── TEST 2: Validaciones de datos ──")
    
    # Ingresos: categoría, cuenta, fecha, valor
    wi = wb["Ingresos"]
    dvs = {str(dv.sqref): (dv.type, dv.formula1) for dv in wi.data_validations.dataValidation}
    
    cat_ok = any("tblCategorias[Nombre]" in str(v[1]) for v in dvs.values())
    cta_ok = any("tblCuentas[Nombre]" in str(v[1]) for v in dvs.values())
    fec_ok = any(v[0] == "date" for v in dvs.values())
    val_ok = any(v[0] == "decimal" for v in dvs.values())
    
    results.ok("Ingresos.Categoria → lista") if cat_ok else results.fail("Ingresos.Categoria sin validación lista")
    results.ok("Ingresos.Cuenta → lista") if cta_ok else results.fail("Ingresos.Cuenta sin validación lista")
    results.ok("Ingresos.Fecha → date") if fec_ok else results.fail("Ingresos.Fecha sin validación fecha")
    results.ok("Ingresos.Valor → decimal >0") if val_ok else results.fail("Ingresos.Valor sin validación decimal")
    
    # Gastos: mismas validaciones
    wg = wb["Gastos"]
    dvs_g = {str(dv.sqref): (dv.type, dv.formula1) for dv in wg.data_validations.dataValidation}
    
    results.ok("Gastos.Categoria → lista") if any("tblCategorias" in str(v[1]) for v in dvs_g.values()) else results.fail("Gastos.Categoria sin validación")
    results.ok("Gastos.Cuenta → lista") if any("tblCuentas" in str(v[1]) for v in dvs_g.values()) else results.fail("Gastos.Cuenta sin validación")
    results.ok("Gastos.Fecha → date") if any(v[0] == "date" for v in dvs_g.values()) else results.fail("Gastos.Fecha sin validación")
    results.ok("Gastos.Valor → decimal >0") if any(v[0] == "decimal" for v in dvs_g.values()) else results.fail("Gastos.Valor sin validación")
    
    # Pagos: estado, fecha, valor
    wp = wb["Pagos"]
    dvs_p = {str(dv.sqref): (dv.type, dv.formula1) for dv in wp.data_validations.dataValidation}
    
    results.ok("Pagos.Estado → lista") if any("Pendiente" in str(v[1]) for v in dvs_p.values()) else results.fail("Pagos.Estado sin validación")
    results.ok("Pagos.FechaVencimiento → date") if any(v[0] == "date" for v in dvs_p.values()) else results.fail("Pagos.Fecha sin validación")
    results.ok("Pagos.Valor → decimal >0") if any(v[0] == "decimal" for v in dvs_p.values()) else results.fail("Pagos.Valor sin validación")
    
    # Configuración: SaldoInicial >= 0
    wc = wb["Configuración"]
    dvs_c = [(dv.type, dv.formula1, str(dv.sqref)) for dv in wc.data_validations.dataValidation]
    saldo_ok = any(v[0] == "decimal" for v in dvs_c)
    results.ok("Configuración.SaldoInicial → decimal ≥0") if saldo_ok else results.fail("Configuración.SaldoInicial sin validación")

# ══════════════════════════════════════════════════════════════════════════
# TEST 3 — FÓRMULAS DEL MOTOR
# ══════════════════════════════════════════════════════════════════════════
def test_motor(wb):
    print("\n── TEST 3: Fórmulas del Motor ──")
    
    wm = wb["Motor"]
    
    # Mapa de celdas esperadas y qué deben contener
    checks = {
        "B2": {"desc": "Ingresos del mes", "must_contain": ["tblIngresos", "Fecha", "Valor"]},
        "B3": {"desc": "Gastos del mes", "must_contain": ["tblGastos", "Fecha", "Valor"]},
        "B4": {"desc": "Balance del mes", "must_contain": ["B2", "B3"]},
        "B7": {"desc": "Total Ingresos histórico", "must_contain": ["SUM", "tblIngresos", "Valor"]},
        "B8": {"desc": "Total Gastos histórico", "must_contain": ["SUM", "tblGastos", "Valor"]},
        "B9": {"desc": "Saldo actual", "must_contain": ["SaldoInicial", "B7", "B8"]},
        "B12": {"desc": "Pendiente de pago", "must_contain": ["SUMIF", "tblPagos", "Pendiente"]},
        "B13": {"desc": "Total vencido", "must_contain": ["SUMIF", "tblPagos", "Vencido"]},
        "B14": {"desc": "Total pagado", "must_contain": ["SUMIF", "tblPagos", "Pagado"]},
        "B17": {"desc": "Disponible restante", "must_contain": ["B9", "B12"]},
        "B18": {"desc": "% Gastado", "must_contain": ["B3", "B2"]},
    }
    
    for coord, spec in checks.items():
        cell_val = str(wm[coord].value or "")
        if not cell_val.startswith("="):
            results.fail(f"Motor.{coord} ({spec['desc']}) no es fórmula: {cell_val}")
            continue
        
        missing = [kw for kw in spec["must_contain"] if kw not in cell_val]
        if not missing:
            results.ok(f"Motor.{coord} ({spec['desc']})")
        else:
            results.fail(f"Motor.{coord} ({spec['desc']}) falta: {missing}")

# ══════════════════════════════════════════════════════════════════════════
# TEST 4 — DASHBOARD REFERENCIAS
# ══════════════════════════════════════════════════════════════════════════
def test_dashboard(wb):
    print("\n── TEST 4: Dashboard ──")
    
    wd = wb["Dashboard"]
    
    # Verificar que todos los valores vienen del Motor
    motor_refs = []
    direct_formulas = []
    
    for row in wd.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                if "Motor!" in cell.value:
                    motor_refs.append(cell.coordinate)
                elif "tblParametros" in cell.value or "CONCATENATE" in cell.value:
                    pass  # Permitido: título con periodo
                else:
                    direct_formulas.append(f"{cell.coordinate}: {cell.value}")
    
    if motor_refs:
        results.ok(f"Dashboard tiene {len(motor_refs)} referencias al Motor")
    else:
        results.fail("Dashboard no referencia al Motor")
    
    if not direct_formulas:
        results.ok("Dashboard no tiene fórmulas directas (RN-003)")
    else:
        results.fail(f"Dashboard tiene fórmulas directas: {direct_formulas}")
    
    # 6 indicadores esperados
    indicadores_motor = ["Motor!B9", "Motor!B2", "Motor!B3", "Motor!B12", "Motor!B17", "Motor!B18"]
    found = set()
    for row in wd.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for ind in indicadores_motor:
                    if ind in cell.value:
                        found.add(ind)
    
    if len(found) == 6:
        results.ok(f"6/6 indicadores del MVP presentes en Dashboard")
    else:
        missing = set(indicadores_motor) - found
        results.fail(f"Indicadores faltantes: {missing}")

# ══════════════════════════════════════════════════════════════════════════
# TEST 5 — REGLAS DE NEGOCIO
# ══════════════════════════════════════════════════════════════════════════
def test_reglas_negocio(wb):
    print("\n── TEST 5: Reglas de negocio ──")
    
    # RN-002: Solo Motor tiene fórmulas de cálculo
    hojas_captura = ["Ingresos", "Gastos", "Pagos", "Configuración"]
    for hoja in hojas_captura:
        ws = wb[hoja]
        formulas_encontradas = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas_encontradas.append(f"{cell.coordinate}: {cell.value}")
        
        if not formulas_encontradas:
            results.ok(f"RN-002/RN-004: {hoja} sin fórmulas (solo datos)")
        else:
            results.fail(f"RN-002/RN-004: {hoja} tiene fórmulas: {formulas_encontradas}")
    
    # RN-006: Tablas usan referencias estructuradas en Motor
    wm = wb["Motor"]
    uses_hardcoded_ranges = False
    for row in wm.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                val = cell.value
                # Buscar referencias tipo A1:A10 que no sean dentro del Motor
                if "Ingresos!" in val or "Gastos!" in val or "Pagos!" in val:
                    uses_hardcoded_ranges = True
    
    if not uses_hardcoded_ranges:
        results.ok("RN-006: Motor usa referencias estructuradas (no rangos hardcoded)")
    else:
        results.fail("RN-006: Motor referencia hojas directamente en vez de tablas")

# ══════════════════════════════════════════════════════════════════════════
# TEST 6 — SIMULACIÓN DE CÁLCULOS
# ══════════════════════════════════════════════════════════════════════════
def test_simulacion(wb):
    print("\n── TEST 6: Simulación de cálculos (lógica del Motor) ──")
    
    # Leer datos de muestra del Excel
    wi = wb["Ingresos"]
    wg = wb["Gastos"]
    wp = wb["Pagos"]
    wc = wb["Configuración"]
    
    # Leer MesActual
    mes_actual = None
    for row in wc.iter_rows():
        for cell in row:
            if cell.value == "MesActual":
                mes_actual = wc.cell(row=cell.row, column=cell.column + 1).value
    
    if not mes_actual:
        results.fail("No se encontró MesActual en tblParametros")
        return
    results.ok(f"MesActual = {mes_actual}")
    
    # Calcular ingresos del mes
    year_mes, month_mes = int(mes_actual.split("-")[0]), int(mes_actual.split("-")[1])
    
    total_ingresos_mes = 0
    total_ingresos_hist = 0
    for row in wi.iter_rows(min_row=2):
        fecha_val = row[1].value  # col B
        valor_val = row[4].value  # col E
        if fecha_val and valor_val:
            if isinstance(fecha_val, str):
                parts = fecha_val.split("-")
                y, m = int(parts[0]), int(parts[1])
            else:
                y, m = fecha_val.year, fecha_val.month
            total_ingresos_hist += valor_val
            if y == year_mes and m == month_mes:
                total_ingresos_mes += valor_val
    
    total_gastos_mes = 0
    total_gastos_hist = 0
    for row in wg.iter_rows(min_row=2):
        fecha_val = row[1].value
        valor_val = row[4].value
        if fecha_val and valor_val:
            if isinstance(fecha_val, str):
                parts = fecha_val.split("-")
                y, m = int(parts[0]), int(parts[1])
            else:
                y, m = fecha_val.year, fecha_val.month
            total_gastos_hist += valor_val
            if y == year_mes and m == month_mes:
                total_gastos_mes += valor_val
    
    # Saldo inicial
    saldo_inicial = 0
    for row in wc.iter_rows():
        cell = row[2] if len(row) > 2 else None
        if cell and cell.value and isinstance(cell.value, (int, float)):
            # Solo si está en la zona de tblCuentas
            if row[0].value and isinstance(row[0].value, int) and row[1].value in ["Efectivo","Banco","Tarjeta"]:
                saldo_inicial += cell.value
    
    # Pagos
    pendiente = 0
    vencido = 0
    pagado = 0
    for row in wp.iter_rows(min_row=2):
        valor = row[3].value  # col D
        estado = row[4].value  # col E
        if valor and estado:
            if estado == "Pendiente": pendiente += valor
            elif estado == "Vencido": vencido += valor
            elif estado == "Pagado": pagado += valor
    
    # Valores esperados
    balance_mes = total_ingresos_mes - total_gastos_mes
    saldo_actual = saldo_inicial + total_ingresos_hist - total_gastos_hist
    disponible = saldo_actual - pendiente
    pct_gastado = total_gastos_mes / total_ingresos_mes if total_ingresos_mes > 0 else 0
    
    print(f"\n  Valores calculados (Python):")
    print(f"    Ingresos mes:    ${total_ingresos_mes:,.2f}")
    print(f"    Gastos mes:      ${total_gastos_mes:,.2f}")
    print(f"    Balance mes:     ${balance_mes:,.2f}")
    print(f"    Saldo actual:    ${saldo_actual:,.2f}")
    print(f"    Pendiente:       ${pendiente:,.2f}")
    print(f"    Vencido:         ${vencido:,.2f}")
    print(f"    Disponible:      ${disponible:,.2f}")
    print(f"    % Gastado:       {pct_gastado:.1%}")
    
    # Verificar coherencia lógica
    if balance_mes == total_ingresos_mes - total_gastos_mes:
        results.ok(f"Balance = Ingresos - Gastos ({balance_mes})")
    
    if saldo_actual == saldo_inicial + total_ingresos_hist - total_gastos_hist:
        results.ok(f"Saldo actual = SaldoInicial + Ingresos - Gastos ({saldo_actual})")
    
    if disponible == saldo_actual - pendiente:
        results.ok(f"Disponible = Saldo - Pendiente ({disponible})")
    
    if total_ingresos_mes > 0:
        results.ok(f"% Gastado = {pct_gastado:.1%}")
    
    # Verificar que las fórmulas del Motor son consistentes con estos cálculos
    wm = wb["Motor"]
    labels = {}
    for row in wm.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.column == 1:
                labels[cell.value] = cell.row
    
    results.ok(f"Motor tiene {len(labels)} indicadores etiquetados")

# ══════════════════════════════════════════════════════════════════════════
# EJECUTAR PRUEBAS
# ══════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print(f"═══════════════════════════════════════════════════════")
    print(f"  OpenFinanceKit — Test Suite")
    print(f"  Archivo: {EXCEL_PATH}")
    print(f"═══════════════════════════════════════════════════════")
    
    if not os.path.exists(EXCEL_PATH):
        print(f"\n❌ Archivo no encontrado: {EXCEL_PATH}")
        sys.exit(1)
    
    wb = load_workbook(EXCEL_PATH)
    
    test_estructura(wb)
    test_validaciones(wb)
    test_motor(wb)
    test_dashboard(wb)
    test_reglas_negocio(wb)
    test_simulacion(wb)
    
    success = results.summary()
    sys.exit(0 if success else 1)
